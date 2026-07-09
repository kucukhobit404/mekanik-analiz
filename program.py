import streamlit as st
import pandas as pd
import openpyxl
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import plotly.express as px

# Sayfa Genişlik Ayarı ve Başlık (Her zaman ilk sırada olmalıdır)
st.set_page_config(page_title="Mekanik İşler Fiyat Analiz Paneli", layout="wide")
st.title("🔧⚙️ Mekanik İşler Fiyat Analiz Paneli")
st.markdown("---")

# ==========================================
# CANLI TCMB KUR SERVİSİ
# ==========================================
@st.cache_data(ttl=3600)
def tcmb_kur_getir(tarih_str, doviz_cinsi="USD"):
    try:
        dt = datetime.strptime(str(tarih_str).strip(), "%d.%m.%Y")
    except:
        return None
    
    for _ in range(7):
        if dt.weekday() == 5:
            dt -= timedelta(days=1)
        elif dt.weekday() == 6:
            dt -= timedelta(days=2)
            
        yil_ay = dt.strftime("%Y%m")
        gun_ay_yil = dt.strftime("%d%m%Y")
        url = f"https://www.tcmb.gov.tr/kurlar/{yil_ay}/{gun_ay_yil}.xml"
        
        try:
            res = requests.get(url, timeout=2)
            if res.status_code == 200:
                root = ET.fromstring(res.content)
                for curr in root.findall('Currency'):
                    if curr.get('CurrencyCode') == doviz_cinsi:
                        val = curr.find('ForexSelling').text
                        if val:
                            return float(val)
        except:
            pass
        dt -= timedelta(days=1)
    return None

@st.cache_data
def sayfalari_getir(dosya_yolu):
    try:
        return pd.ExcelFile(dosya_yolu).sheet_names
    except:
        return None

# ==========================================
# AKILLI VERİ YÜKLEME VE SÜTUN EŞİTLEYİCİ
# ==========================================
@st.cache_data
def veri_yukle(dosya_yolu, sayfa_adi):
    try:
        df = pd.read_excel(dosya_yolu, sheet_name=sayfa_adi, header=None)
        header_row_idx = None
        
        anahtar_kelimeler = ["Yapılacak İşin Cinsi", "Tarih", "Fiyat", "Marka", "TÜR", "Sıra No"]
        
        for idx, row in df.iterrows():
            if row.astype(str).str.contains("|".join(anahtar_kelimeler), case=False, na=False).any():
                header_row_idx = idx
                break
        
        if header_row_idx is not None:
            raw_headers = df.iloc[header_row_idx].tolist()
            num_df_cols = len(df.columns)
            cleaned_headers = []
            
            for i in range(num_df_cols):
                val = raw_headers[i] if i < len(raw_headers) else None
                if pd.notna(val) and str(val).strip() != "":
                    col_name = str(val).strip()
                    if col_name in cleaned_headers:
                        col_name = f"{col_name}_{i}"
                    cleaned_headers.append(col_name)
                else:
                    cleaned_headers.append(f"Boş_Sütun_{i+1}")
            
            df_temiz = df.iloc[header_row_idx + 1:].copy()
            df_temiz.columns = cleaned_headers
            df = df_temiz
        else:
            df = pd.read_excel(dosya_yolu, sheet_name=sayfa_adi)
            df.columns = [str(c).strip() for c in df.columns]
        
        if 'Tarih' in df.columns:
            df['Tarih'] = pd.to_datetime(df['Tarih'], errors='coerce').dt.strftime('%d.%m.%Y')
            
        df = df.loc[:, ~df.columns.str.contains('Boş_Sütun')]
        return df
    except Exception as e:
        st.error(f"Veri yükleme hatası: {e}")
        return pd.DataFrame()

def excel_satir_ekle(dosya_yolu, sayfa_adi, yeni_veri_dict):
    try:
        wb = openpyxl.load_workbook(dosya_yolu)
        ws = wb[sayfa_adi]
        header_row_idx = None
        
        anahtar_kelimeler = ["Yapılacak İşin Cinsi", "Tarih", "Fiyat", "Marka", "TÜR", "Sıra No"]
        for r in range(1, 20):
            satir_degerleri = [str(ws.cell(row=r, column=c).value) for c in range(1, 20)]
            if any(any(k.lower() in val.lower() for k in anahtar_kelimeler) for val in satir_degerleri if val):
                header_row_idx = r
                break
        
        if not header_row_idx:
            return False
            
        kolon_haritasi = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row_idx, column=col_idx).value
            if cell_val:
                kolon_haritasi[str(cell_val).strip()] = col_idx
        
        yeni_satir_idx = ws.max_row + 1
        for kolon_adi, deger in yeni_veri_dict.items():
            if kolon_adi in kolon_haritasi:
                target_col = kolon_haritasi[kolon_adi]
                if str(deger).isdigit():
                    deger = int(deger)
                else:
                    try:
                        deger = float(str(deger).replace(',', '.'))
                    except:
                        pass
                ws.cell(row=yeni_satir_idx, column=target_col, value=deger)
        
        wb.save(dosya_yolu)
        wb.close()
        return True
    except Exception as e:
        st.error(f"Excel'e yazarken hata oluştu: {e}")
        return False

# ==========================================
# YENİ MODÜL: VANA & FLANŞ VERİTABANI YÜKLEME
# ==========================================
@st.cache_data
def vana_flans_veritabani_yukle():
    try:
        # Excel dosyanızın adı (Aynı klasörde olduğundan emin olun)
        dosya_adi = "Çalışma.xlsx" 
        
        # 'sheet_name' kısımlarını Excel'inizdeki sekme isimleriyle aynı yapın
        # (Önemli: Excel dosyanızı açıp sekmelerin adının tam olarak "Vanalar" ve "Paslanmaz Fittings" olduğunu kontrol edin)
        vanalar_db = pd.read_excel(dosya_adi, sheet_name="Vanalar")
        fittings_db = pd.read_excel(dosya_adi, sheet_name="Paslanmaz Fittings")
        
        # Sütun isimlerindeki boşlukları temizleme (Hata almamak için)
        vanalar_db.columns = vanalar_db.columns.str.strip()
        fittings_db.columns = fittings_db.columns.str.strip()
        
        return vanalar_db, fittings_db
    except Exception as e:
        st.error(f"HATA: Excel dosyası bulunamadı veya sekme isimleri yanlış! Hata detayı: {e}")
        return pd.DataFrame(), pd.DataFrame()

dosya_yolu = "fiyat_veritabani.xlsx"
sayfalar = sayfalari_getir(dosya_yolu)

if sayfalar is None:
    st.error("Excel dosyası okunamadı! Klasörde 'fiyat_veritabani.xlsx' dosyasının olduğundan emin olun.")
else:
    # ==========================================
    # MODÜLER MİMARİ VE EKRAN İZOLASYONU
    # ==========================================
    # 🖥️ Ekran Mimarisi başlığını sildik, yerine logoyu ekledik
    st.sidebar.image("logo.jpg", use_container_width=True) 
    
    # Görselden sonra temiz bir ayrım için boşluk veya çizgi ekleyebiliriz
    st.sidebar.markdown("---") 

    # 4. Modül buraya eklendi
    ana_ekran = st.sidebar.radio(
        "Çalışma Alanı Seçimi:", 
        [
            "🔍 Yönetim & Analiz Paneli", 
            "➕ Yeni Veri Girişi", 
            "🏗️ Çelik Ağırlık Hesabı",
            "🚰 Vana ve Flanş Seçim Aracı" 
        ],
        key="ana_menu_secimi" # <-- BU KEY PARAMETRESİNİ MUTLAKA EKLEYİN
    )
    
    # Sol Sidebar Genel Ayarları
    st.sidebar.header("📁 Veritabanı Grubu")
    secilen_sekme = st.sidebar.selectbox("İşlem Yapılacak Grup:", sayfalar)
    
    st.sidebar.markdown("---")
    st.sidebar.header("💱 Döviz Kur Ayarları")
    doviz_secimi = st.sidebar.selectbox("Tablo Döviz Cinsi:", ["USD", "EUR"])
    
    bugun_str = datetime.now().strftime("%d.%m.%Y")
    canli_bugun_kuru = tcmb_kur_getir(bugun_str, doviz_secimi)
    if not canli_bugun_kuru:
        canli_bugun_kuru = 34.50
        
    bugunku_kur_girdisi = st.sidebar.number_input(
        f"Bugünkü Canlı {doviz_secimi}/TL Kuru:", 
        value=float(canli_bugun_kuru), 
        step=0.01
    )
    
    # Ana veriyi yükleme
    df = veri_yukle(dosya_yolu, secilen_sekme)

    # ------------------------------------------
    # EKRAN 1: YÖNETİM PANELİ (Sadece Arama ve Grafik)
    # ------------------------------------------
    if ana_ekran == "🔍 Yönetim & Analiz Paneli":
        if not df.empty:
            st.subheader(f"🛠️ {secilen_sekme} Arama Kriterleri Yönetimi")
            
            haric_kolonlar = ["Sıra No", "Fiyat", "Tarih"]
            mevcut_arama_kolonlari = [c for c in df.columns if c not in haric_kolonlar]
            
            varsayilan_kriterler = []
            if "TÜR" in df.columns:
                varsayilan_kriterler = [c for c in ["Yapılacak İşin Cinsi", "TÜR", "Marka"] if c in df.columns]
            else:
                varsayilan_kriterler = mevcut_arama_kolonlari[:3] if len(mevcut_arama_kolonlari) >= 3 else mevcut_arama_kolonlari

            secilen_filtre_kolonlari = st.multiselect(
                "🔍 Ekranda görmek istediğiniz arama/filtre sütunlarını seçin veya kaldırın:",
                options=mevcut_arama_kolonlari,
                default=varsayilan_kriterler
            )
            
            filtrelenmis_df = df.copy()
            secim_yapildi_mi = False
            
            if secilen_filtre_kolonlari:
                st.markdown("##### 🎛️ Seçtiğiniz Kriterlere Göre Filtreleyin:")
                arama_kolonlari = st.columns(len(secilen_filtre_kolonlari))
                
                for i, sutun in enumerate(secilen_filtre_kolonlari):
                    with arama_kolonlari[i]:
                        mevcut_secenekler = ["Seçiniz..."] + list(filtrelenmis_df[sutun].dropna().astype(str).unique())
                        secim = st.selectbox(f"{sutun}", mevcut_secenekler, key=f"Arama_{secilen_sekme}_{sutun}")
                        
                        if secim != "Seçiniz...":
                            filtrelenmis_df = filtrelenmis_df[filtrelenmis_df[sutun].astype(str) == secim]
                            secim_yapildi_mi = True
            else:
                st.info("💡 Yukarıdaki kutudan filtreleme kriterleri seçerek aramayı daraltabilirsiniz.")

            st.markdown("---")
            if secim_yapildi_mi:
                st.subheader(f"📊 Sonuçlar ({len(filtrelenmis_df)} Ürün Listelendi)")
                st.dataframe(filtrelenmis_df, use_container_width=True, hide_index=True)
                
                if 'Fiyat' in filtrelenmis_df.columns and 'Tarih' in filtrelenmis_df.columns:
                    st.markdown("### 📈 Keşif Ön Analizi ve Kur Değişim Grafiği")
                    
                    chart_data = []
                    yil_bazli_maliyet = {}
                    
                    with st.spinner("TCMB Geçmiş Tarihli Kurlar Sorgulanıyor..."):
                        for idx, row in filtrelenmis_df.iterrows():
                            raw_fiyat = str(row['Fiyat']).replace(r'[^\d.]', '')
                            fiyat_val = pd.to_numeric(raw_fiyat, errors='coerce')
                            tarih_val = str(row['Tarih']).strip()
                            
                            if pd.notna(fiyat_val) and tarih_val:
                                o_gunk_kur = tcmb_kur_getir(tarih_val, doviz_cinsi=doviz_secimi)
                                if not o_gunk_kur: 
                                    o_gunk_kur = bugunku_kur_girdisi
                                
                                o_gunk_tl = fiyat_val * o_gunk_kur
                                bugunk_tl = fiyat_val * bugunku_kur_girdisi
                                
                                try:
                                    yil_bilgisi = tarih_val.split('.')[-1]
                                    if len(yil_bilgisi) != 4 or not yil_bilgisi.isdigit():
                                        yil_bilgisi = "Bilinmeyen Yıl"
                                except:
                                    yil_bilgisi = "Bilinmeyen Yıl"
                                    
                                if yil_bilgisi not in yil_bazli_maliyet:
                                    yil_bazli_maliyet[yil_bilgisi] = {"o_gunk_toplam": 0.0, "bugunk_toplam": 0.0, "adet": 0}
                                
                                yil_bazli_maliyet[yil_bilgisi]["o_gunk_toplam"] += o_gunk_tl
                                yil_bazli_maliyet[yil_bilgisi]["bugunk_toplam"] += bugunk_tl
                                yil_bazli_maliyet[yil_bilgisi]["adet"] += 1
                                
                                if 'Sıra No' in row and pd.notna(row['Sıra No']):
                                    etiket = f"Sıra {row['Sıra No']}"
                                else:
                                    etiket = f"Ürün {len(chart_data) + 1}"
                                    
                                if 'Marka' in row and pd.notna(row['Marka']):
                                    etiket += f" ({row['Marka']})"
                                elif 'TÜR' in row and pd.notna(row['TÜR']):
                                    etiket += f" ({row['TÜR']})"
                                    
                                chart_data.append({
                                    "Ürün Tanımı": etiket,
                                    "O Günkü Kurlu Değer (TL)": round(o_gunk_tl, 2),
                                    "Bugünkü Kurlu Güncel Değer (TL)": round(bugunk_tl, 2),
                                    "Yıl": yil_bilgisi
                                })
                    
                    if chart_data:
                        for yil in sorted(yil_bazli_maliyet.keys()):
                            veriler = yil_bazli_maliyet[yil]
                            ortalama_o_gunk_tl = veriler["o_gunk_toplam"] / veriler["adet"]
                            ortalama_bugunk_tl = veriler["bugunk_toplam"] / veriler["adet"]
                            fark_oran = ((ortalama_bugunk_tl - ortalama_o_gunk_tl) / ortalama_o_gunk_tl * 100) if ortalama_o_gunk_tl > 0 else 0
                            
                            st.markdown(f"#### 📅 {yil} Yılı Hesaplanan Ortalamalar ({veriler['adet']} Ürün)")
                            m1, m2, m3 = st.columns(3)
                            m1.metric("O Günkü Kura Göre Ortalama", f"{ortalama_o_gunk_tl:,.2f} TL")
                            m2.metric("Bugünkü Kura Göre Ortalama", f"{ortalama_bugunk_tl:,.2f} TL")
                            m3.metric("Kur Değişim Etkisi (Fark)", f"{fark_oran:+.2f} %")
                            st.markdown("---")
                        
                        df_chart = pd.DataFrame(chart_data)
                        df_chart = df_chart.sort_values(by=["Yıl", "Ürün Tanımı"])
                        
                        st.write("🔍 **Yıllara Göre Gruplandırılmış Ürün Maliyet Karşılaştırması:**")
                        
                        fig = px.bar(
                            df_chart, 
                            x="Ürün Tanımı", 
                            y=["O Günkü Kurlu Değer (TL)", "Bugünkü Kurlu Güncel Değer (TL)"],
                            barmode="group",
                            facet_col="Yıl",
                            facet_col_wrap=2, 
                            color_discrete_sequence=["#00C9FF", "#FF9A44"], 
                            labels={"value": "Maliyet (TL)", "variable": "Kur Durumu"}
                        )

                        fig.for_each_annotation(lambda a: a.update(text=f"📊 {a.text.split('=')[-1]} Yılı Analizi"))

                        fig.update_layout(
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            margin=dict(l=20, r=20, t=40, b=20),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                        )

                        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(128, 128, 128, 0.2)')
                        fig.update_xaxes(matches=None, showticklabels=True)

                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        st.warning("Seçilen ürünlerin fiyat veya tarih bilgisi grafik oluşturmak için uygun formatta değil.")
            else:
                st.info("💡 Canlı kur analiz grafiklerini görmek için yukarıdaki kriterlerden seçim yapın.")
                if st.checkbox("Tüm Listeyi Göster (Filtresiz)"):
                    st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.warning("Veri bulunamadı.")

    # ------------------------------------------
    # EKRAN 2: YENİ VERİ GİRİŞİ (Bağımsız Modül)
    # ------------------------------------------
    elif ana_ekran == "➕ Yeni Veri Girişi":
        st.subheader(f"📝 {secilen_sekme} Kategorisine Yeni Ürün Ekle")
        if not df.empty:
            girdi_verileri = {}
            sira_no_mevcut = "Sıra No" in df.columns
            
            if sira_no_mevcut:
                try:
                    son_sira = pd.to_numeric(df["Sıra No"], errors='coerce').max()
                    yeni_sira_no = int(son_sira) + 1 if pd.notna(son_sira) else 1
                except:
                    yeni_sira_no = len(df) + 1
                
                st.info(f"ℹ️ Bu ürün için otomatik atanacak **Sıra No: {yeni_sira_no}**")
                girdi_verileri["Sıra No"] = yeni_sira_no
                form_kolonlari = [c for c in df.columns if c != "Sıra No"]
            else:
                form_kolonlari = list(df.columns)
            
            with st.form("yeni_ekipman_formu", clear_on_submit=True):
                form_grid = st.columns(2)
                for idx, kolon in enumerate(form_kolonlari):
                    target_grid = form_grid[idx % 2]
                    with target_grid:
                        if kolon == "Tarih":
                            bugun_tarihi = datetime.now().strftime("%d.%m.%Y")
                            girdi_verileri[kolon] = st.text_input(f"{kolon}:", value=bugun_tarihi)
                        else:
                            girdi_verileri[kolon] = st.text_input(f"{kolon}:")
                
                st.markdown(" ")
                kaydet_butonu = st.form_submit_button("⚡ Ürünü Veritabanına Güvenle Kaydet")
                
                if kaydet_butonu:
                    if form_kolonlari and girdi_verileri[form_kolonlari[0]].strip() == "":
                        st.error(f"Lütfen en azından '{form_kolonlari[0]}' alanını doldurun!")
                    else:
                        basarili_mi = excel_satir_ekle(dosya_yolu, secilen_sekme, girdi_verileri)
                        if basarili_mi:
                            st.success(f"🎉 Ürün başarıyla '{secilen_sekme}' tablosuna eklendi!")
                            st.cache_data.clear()
                            st.rerun()
        else:
            st.warning("Veri bulunamadı. Lütfen sol menüden geçerli bir veritabanı grubu seçin.")

    # ------------------------------------------
    # EKRAN 3: ÇELİK AĞIRLIK HESABI 
    # ------------------------------------------
    elif ana_ekran == "🏗️ Çelik Ağırlık Hesabı":
        st.subheader("🏗️ Çelik Ağırlık Hesabı & İmalat Maliyet Modülü")
        st.info("Çelik ağırlık hesabı ve metraj modundasınız. Yönetim paneli, ham veri tabloları ve sepet mekanizmaları gizlenmiştir.")
        
        # Detaylandırılmış Geniş Mühendislik Birim Ağırlık Veritabanı
        profil_veritabani = {
            "HEA": {
                "HEA 100": 16.7, "HEA 120": 19.9, "HEA 140": 24.7, "HEA 160": 30.4, 
                "HEA 180": 35.5, "HEA 200": 42.3, "HEA 220": 50.5, "HEA 240": 60.3, 
                "HEA 260": 68.2, "HEA 280": 76.4, "HEA 300": 88.3
            },
            "IPE": {
                "IPE 100": 8.1, "IPE 120": 10.4, "IPE 140": 12.9, "IPE 160": 15.8, 
                "IPE 180": 18.8, "IPE 200": 22.4, "IPE 220": 26.2, "IPE 240": 30.7, 
                "IPE 270": 36.1, "IPE 300": 42.2, "IPE 330": 49.1, "IPE 360": 57.1
            },
            "NPU": {
                "NPU 65": 7.09, "NPU 80": 8.64, "NPU 100": 10.60, "NPU 120": 13.40, 
                "NPU 140": 16.00, "NPU 160": 18.80, "NPU 180": 22.00, "NPU 200": 25.30,
                "NPU 220": 29.40, "NPU 240": 33.20, "NPU 300": 46.20
            },
            "Köşebent (L Profil)": {
                "30x30x3 mm": 1.36, "40x40x4 mm": 2.42, "50x50x5 mm": 3.77, 
                "60x60x6 mm": 5.42, "70x70x7 mm": 7.38, "80x80x8 mm": 9.66, 
                "100x100x10 mm": 15.10, "120x120x12 mm": 21.60
            },
            "Lama (Düz Bar)": {
                "20x5 mm": 0.79, "30x5 mm": 1.18, "40x5 mm": 1.57, 
                "50x10 mm": 3.93, "60x10 mm": 4.71, "80x10 mm": 6.28, 
                "100x10 mm": 7.85, "120x12 mm": 11.30, "150x15 mm": 17.66
            },
            "Kare Profil": {
                "20x20x2 mm": 1.12, "30x30x2 mm": 1.75, "40x40x2 mm": 2.38, 
                "50x50x3 mm": 4.28, "60x60x3 mm": 5.22, "80x80x4 mm": 9.22, 
                "100x100x4 mm": 11.73, "120x120x5 mm": 17.40, "150x150x6 mm": 26.50
            },
            "Dikdörtgen Profil": {
                "30x20x2 mm": 1.44, "40x20x2 mm": 1.75, "50x30x2 mm": 2.38, 
                "60x40x3 mm": 4.28, "80x40x3 mm": 5.22, "100x50x4 mm": 8.59, 
                "120x60x4 mm": 10.70, "150x100x5 mm": 18.50, "200x100x6 mm": 26.50
            }
        }
        
        m1, m2, m3 = st.columns(3)
        with m1:
            profil_serisi = st.selectbox("Çelik Profil Grubu:", list(profil_veritabani.keys()))
        with m2:
            secenekler = list(profil_veritabani[profil_serisi].keys())
            secilen_kesit = st.selectbox("Profil Boyutu / Piyasa Ölçüsü:", secenekler)
            birim_agirlik = profil_veritabani[profil_serisi][secilen_kesit]
        with m3:
            malzeme_yapisi = st.selectbox("Malzeme Bileşeni:", ["Karbon Çeliği", "AISI 304 Paslanmaz", "AISI 316 Paslanmaz"])
            
        col_metraj, col_endeks = st.columns(2)
        with col_metraj:
            toplam_metre = st.number_input("Toplam Metraj (Metre):", value=12.0, step=1.0, min_value=0.0)
        with col_endeks:
            default_endeks = 1.20 if malzeme_yapisi == "Karbon Çeliği" else (3.80 if malzeme_yapisi == "AISI 304 Paslanmaz" else 5.20)
            hammadde_endeksi = st.number_input(f"Hammadde Maliyet Endeksi ({doviz_secimi}/kg):", value=default_endeks, step=0.1)
            
        # Yapısal Mühendislik Hesaplamaları
        toplam_agirlik_kg = birim_agirlik * toplam_metre
        toplam_tonaj = toplam_agirlik_kg / 1000.0
        toplam_maliyet_doviz = toplam_agirlik_kg * hammadde_endeksi
        toplam_maliyet_tl = toplam_maliyet_doviz * bugunku_kur_girdisi
        
        st.markdown("---")
        st.markdown("#### 📈 Analiz Sonuçları")
        
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Profil Birim Ağırlığı", f"{birim_agirlik} kg/m")
        c2.metric("Toplam Hesaplanan Ağırlık", f"{toplam_agirlik_kg:,.2f} kg")
        c3.metric("Toplam Tonaj", f"{toplam_tonaj:,.4f} Ton")
        c4.metric("Seçilen Profil Detayı", str(secilen_kesit))
        
        st.markdown("##### 💵 Tahmini Toplam İmalat Maliyeti")
        res_c1, res_c2 = st.columns(2)
        res_c1.metric(f"Maliyet ({doviz_secimi})", f"{toplam_maliyet_doviz:,.2f} {doviz_secimi}")
        res_c2.metric("Maliyet (TL)", f"{toplam_maliyet_tl:,.2f} TL")

    # ------------------------------------------
    # EKRAN 4: VANA VE FLANŞ SEÇİM ARACI (YENİ EKLENDİ)
    # ------------------------------------------
    elif ana_ekran == "🚰 Vana ve Flanş Seçim Aracı":
        st.subheader("🚰 Vana ve Flanş Montaj Seti Seçim Aracı")
        st.info("Bu modül, tesisat hatlarınızdaki vana ve flanş montajları için gerekli olan cıvata, somun, conta setlerini ve flanş ağırlıklarını otomatik hesaplar.")

        vanalar_db, fittings_db = vana_flans_veritabani_yukle()
        
        if vanalar_db.empty or fittings_db.empty:
            st.error("⚠️ 'Çalışma.xlsx - Vanalar.csv' veya 'Çalışma.xlsx - Paslanmaz Fittings.csv' dosyası bulunamadı. Lütfen programın olduğu klasöre ekleyin.")
        else:
            # Dropdown'lar için veritabanındaki eşsiz değerleri (tekrarsız) çekiyoruz
            cap_secenekleri = sorted(vanalar_db['Çap'].dropna().unique().tolist())
            pn_secenekleri = sorted(vanalar_db['Basınç Sınıfı'].dropna().unique().tolist())
            vana_secenekleri = sorted(vanalar_db['Cins'].dropna().unique().tolist())

            # Arayüz Düzeni
            col1, col2, col3 = st.columns(3)
            with col1:
                secilen_cap = st.selectbox("Boru Çapı (DN):", cap_secenekleri)
            with col2:
                secilen_pn = st.selectbox("Basınç Sınıfı:", pn_secenekleri)
            with col3:
                secilen_vana = st.selectbox("Vana Tipi:", vana_secenekleri)

            st.markdown("---")
            
            # Seçimlere göre filtreleme ve sonuç gösterme mantığı
            if st.button("🔧 Montaj Malzemelerini Hesapla", use_container_width=True):
                
                # 1. Vana Filtrelemesi
                vana_filtre = vanalar_db[
                    (vanalar_db['Çap'].astype(str).str.upper() == str(secilen_cap).upper()) &
                    (vanalar_db['Cins'].astype(str).str.lower() == str(secilen_vana).lower()) &
                    (vanalar_db['Basınç Sınıfı'].astype(str).str.upper() == str(secilen_pn).upper())
                ]
                
                # 2. Flanş Filtrelemesi
                flans_filtre = fittings_db[
                    (fittings_db['Çap'].astype(str).str.upper() == str(secilen_cap).upper()) &
                    (fittings_db['Cins'].astype(str).str.lower() == 'flanş') &
                    (fittings_db['Basınç Sınıfı'].astype(str).str.upper() == str(secilen_pn).upper())
                ]
                
                col_sonuc1, col_sonuc2 = st.columns(2)
                
                with col_sonuc1:
                    st.markdown("#### ⚙️ Vana Bağlantı Seti")
                    if not vana_filtre.empty:
                        satir = vana_filtre.iloc[0]
                        delik_sayisi = int(satir['n'])
                        
                        # Kelebek/Wafer vanalarda flanşlar arası tek saplama, diğerlerinde her flanş için ayrı cıvata.
                        if "kelebek" in str(secilen_vana).lower() or "wafer" in str(secilen_vana).lower():
                            civata_adeti = delik_sayisi
                        else:
                            civata_adeti = delik_sayisi * 2
                            
                        civata_olcusu = satir['Civata Çapı']
                        civata_boyu = satir['Cıvata Boyu'] if 'Cıvata Boyu' in satir and pd.notna(satir['Cıvata Boyu']) else "Standart"
                        vana_agirlik = satir['kg/ad'] if 'kg/ad' in satir and pd.notna(satir['kg/ad']) else "-"
                        
                        st.success("✅ Vana kaydı bulundu.")
                        st.metric("Vana Ağırlığı", f"{vana_agirlik} kg/ad")
                        st.metric("Cıvata Ölçüsü", f"{civata_olcusu} (Uzunluk: {civata_boyu})")
                        st.metric("Gerekli Cıvata/Somun Adeti", f"{civata_adeti} Adet")
                        st.metric("Gerekli Conta Adeti", "2 Adet (Sağ/Sol)")
                    else:
                        st.warning(f"Kütüphanede {secilen_cap} - {secilen_pn} ölçülerinde '{secilen_vana}' kaydı bulunamadı.")
                
                with col_sonuc2:
                    st.markdown("#### 🔘 Karşılık Flanş Detayları")
                    if not flans_filtre.empty:
                        f_satir = flans_filtre.iloc[0]
                        flans_kalinlik = f_satir['Kalınlık (mm)']
                        flans_agirlik = f_satir['kg/ad']
                        
                        st.info("ℹ️ Paslanmaz Flanş verileri getirilmiştir.")
                        st.metric("Flanş Delik Sayısı", f"{int(f_satir['n'])} Delikli")
                        st.metric("Flanş Kalınlığı", f"{flans_kalinlik} mm")
                        st.metric("Flanş Birim Ağırlığı", f"{flans_agirlik} kg")
                    else:
                        st.warning(f"Kütüphanede {secilen_cap} - {secilen_pn} paslanmaz flanş kaydı bulunamadı.")